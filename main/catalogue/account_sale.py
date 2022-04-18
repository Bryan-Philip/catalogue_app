from openpyxl import load_workbook
from mysql.connector import connect
from main.catalogue.connector import CONNECTOR
import re
from main.catalogue.interpret import TEAGRADES_DATA
from main.catalogue.query import *
from django.core.files.storage import FileSystemStorage
import json
from main.catalogue.format import *

DATA_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
DATA_ALIAS = {
    "buyer_code": ['buyer'],
    "invoice_number": None,
    "sale_date": None,
    "sale_date_alt": None,
    "prompt_date": None,
    "receiver_address_line1": None,
    "receiver_address_line2": None,
    "receiver_address_line3": None,
    "auction_number": None,
    "lot_number": ['lot no'],
    "invoice_number_buyer": ['invoice'],
    "mark": ['mark'],
    "warehouse": None,
    "packages": ['packages'],
    "type": ['packing type', 'packaging type'],
    "grade": ['grade'],
    "net": ['net weight', 'net weight(in kg)'],
    "gross": ['gross weight', 'gross weight(in kg)'],
    "base_price": ['base price'],
    "broker_starting_price": ['broker starting price', 'broker starting price($)'],
    "price": ['sale price', 'sale price($)'],
    "status": ['status'],
    "sold_packages": ['sold packages'],
    "certification": ['certification'],
    "manufacture_date": ['manf date'],
    "producer": ['producer'],
    "broker": ['broker'],
    "number": ["si no", "sl no"],
    "empty1": None,
    "empty2": None,
}

DATA_ACCOUNT_SALE = {
    "buyer_code": 'K5',
    "account_sale_number": 'K11',
    "sale_date": 'E11',
    "sale_date_alt": 'N7',
    "prompt_date": 'H11',
    "receiver_address_line1": 'A5',
    "receiver_address_line2": 'A6',
    "receiver_address_line3": 'A7',
    "auction_number": 'A9',
    "lot_number": 'A15',
    "invoice_number_buyer": 'B15',
    "packages": 'C15',
    "type": 'D15',
    "grade": 'E15',
    "net": 'F15',
    "price": 'G15',
    "empty1": 'I15',
    "empty2": 'K15',
}

DATA_ACCOUNT_SALE_META = {
    "sale_date_alt": 'N7',
    "receiver_address_line1": 'A5',
    "receiver_address_line2": 'A6',
    "receiver_address_line3": 'A7',
    "auction_number": 'A9',
    "auction_number_alt": 'A11',
    "account_sale_number": 'K11',
    "sale_date": 'E11',
    "prompt_date": 'H11',
}

DATA_ACCOUNT_SALE_TOTALS = {
    'pkgs': 'C',
    'net': 'F',
    'amount': 'H',
    'brokerage': 'J',
    'whtax': 'L',
    'payable_amount': 'M',
}

DATA_ACCOUNT_SALE_TAX_SUMMARY = {
    'amount': 'A',
    'brokerage': 'C',
    'gross': 'E',
    'whtax': 'G',
    'payable_amount': 'I',
}

DATA_SALE_RELATION = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'packages': 'C',
    'type': 'D',
    'grade': 'E',
    'net': 'F',
    'price': 'G',
    '_amount': 'H',
    'empty1': 'I',
    '_brokerage': 'J',
    'empty2': 'K',
    '_whtax': 'L',
    '_payable_amount': 'M',
}

DATA_SALE = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'packages': 'C',
    'type': 'D',
    'grade': 'E',
    'net': 'F',
    'price': 'G',
    '_amount': '=F*G',
    'empty1': 'I',
    '_brokerage': '=H*-0.75%',
    'empty2': 'K',
    '_whtax': '=J*-5%',
    '_payable_amount': '=H+J+L',
}

# Merge M and N

def AccountSaleNumberQuery():
    try:
        with connect(**CONNECTOR) as connection:
            current_number = '''SELECT `number` FROM `account_sale_number`'''
            with connection.cursor() as cursor:
                cursor.execute(current_number)
                row = cursor.fetchone()
                return row[0]

    except Error as e:
            print(e)

ACCOUNT_SALE_COUNTER = AccountSaleNumberQuery()
def InvoiceCounter():
    global ACCOUNT_SALE_COUNTER
    if(ACCOUNT_SALE_COUNTER < 2000):
        ACCOUNT_SALE_COUNTER += 1
        # if len(str(ACCOUNT_SALE_COUNTER)) == 1:
        #     ACCOUNT_SALE_COUNTER = '00' + str(ACCOUNT_SALE_COUNTER)
        # elif len(str(ACCOUNT_SALE_COUNTER)) == 2:
        #     ACCOUNT_SALE_COUNTER = '0' + str(ACCOUNT_SALE_COUNTER)
    else:
        ACCOUNT_SALE_COUNTER = 1
        # ACCOUNT_SALE_COUNTER = '00' + str(1)
    return ACCOUNT_SALE_COUNTER

def CloseAccountSaleNumber(account_sale_number, Pid):
    try:
        with connect(**CONNECTOR) as connection:
            update_lot = '''UPDATE `main_auctions` SET `account_sale_number` = %s WHERE `Pid` = %s'''
            with connection.cursor() as cursor:
                cursor.execute(update_lot, (account_sale_number, Pid,))
                connection.commit()

    except Error as e:
        print(e)

def GenerateAccountSaleNumber(auction_number, auction_number_alt):
    base = 'PRME/'
    base_conform = 'PRME_'
    return {
        'number': base + str(auction_number),
        'number_alt': base + str(auction_number_alt),
        'file': base_conform + str(auction_number)
    }
    # return {
    #     'number': base + str(auction_number) + '/' + str(InvoiceCounter()),
    #     'file': base_conform + str(auction_number) + '_' + str(InvoiceCounter())
    # }

def findAlias(list, value):
    perfect = [
        'grade',
        'mark',
        "number",
    ]
    end = len(list)
    counter = 1
    for alias in list:
        if(value in perfect):
            if(re.search(alias, value) != None):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        else:
            if(value == alias):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        counter += 1
    

def getAliasRelation(value):
    endif = 0
    for alias in DATA_ALIAS:
        endif += 1
        if(DATA_ALIAS[alias] != None):
            if(findAlias(DATA_ALIAS[alias], value.lower())):
                return alias
            else:
                if endif == len(DATA_ALIAS):
                    return alias
                else:
                    continue
        else: continue

class DataInterpretor:
    def init_data(left_bound, data_layer, file):
        left_bound = int(left_bound)
        data_layer = int(data_layer)
        folder='media/documents/sales/'
        fs = FileSystemStorage(location=folder)
        file_data = fs.open(file, 'rb+')
        WORKBOOK = load_workbook(filename = file_data )
        DATA = {}
        RELATION = []
        sheet = WORKBOOK.active
        bc = 0
        inner = list()
        endif_check = list()
        for row in sheet.values:
            allow = True
            inner_counter = 0
            for value in row:
                if(inner_counter >= left_bound-1):
                    if(value != None):
                        if(bc == data_layer-1):
                            value = re.sub(r'\t', '', value)
                            RELATION.append(value)
                        inner.append(value)
                    else:
                        inner.append(None)
                        endif_check.append(value)
                inner_counter += 1
            if(len(endif_check) >= 5 and bc >= 5):
                allow = False
                break
            elif(bc >= data_layer and len(endif_check) < 5):
                if allow: DATA[bc-(data_layer)] = row[slice(left_bound-1, len(row))]
            inner = list()
            endif_check = list()
            bc += 1
        return {
            'relation': RELATION,
            'data': DATA
        }

    def generate_data(data):
        set_data = list()
        counter = 0
        for vals in data['data'].values():
            inner_counter = 0
            inner_dict = dict()
            for inner_data in vals:
                relation = getAliasRelation(data['relation'][inner_counter])
                if(relation != False):
                    inner_dict[relation] = inner_data
                else:
                    inner_dict[relation] = None
                inner_counter += 1
            set_data.append(inner_dict)
            counter += 1
        return set_data

STACK_DATA = {}
combined = list()
LOT_STATUS_RELATION = dict()
populated = list()
WAREHOUSES = list()
WAREHOUSES_RELATION = dict()
WAREHOUSES_OUTLOT = list()  

def StackGenerator(input_data, catalogue_data):
    global STACK_DATA
    global WAREHOUSES
    global WAREHOUSES_RELATION
    global WAREHOUSES_OUTLOT
    counter = 0
    for file in input_data:
        print(file)
        STACK_DATA[counter] = DataInterpretor.generate_data(
            DataInterpretor.init_data(
                file['left_bound'],
                file['data_layer'],
                file['file']
            )
        )
        counter += 1
    STACK_DATA = STACK_DATA[0]
    for lot in STACK_DATA:
        LOT_STATUS_RELATION[lot['lot_number']] = lot['status']
    for data in STACK_DATA:
        exist = list()
        inner_val = dict()
        for single in data:
            if(single in DATA_ACCOUNT_SALE.keys()):
                exist.append(single)
        for value in DATA_ACCOUNT_SALE.keys():
            if(value in exist):
                inner_val[value] = data[value]
            else:
                inner_val[value] = None
        populated.append(inner_val)
    
    folder='media/documents/catalogue_data'
    fsc = FileSystemStorage(location=folder)
    
    with fsc.open(catalogue_data, 'rb+') as fcc_file:
        file_datac = json.load(fcc_file)
        
    for lot in populated:
        lot['warehouse'] = GetInvoiceWarehouse(file_datac, lot['invoice_number_buyer'])
        if lot['invoice_number_buyer'] != '':
            WAREHOUSES.append(re.sub(r'[0-9]', '', GetInvoiceWarehouse(file_datac, lot['invoice_number_buyer'])))
    WAREHOUSES = list(set(WAREHOUSES))
    print(WAREHOUSES)
    for warehouse in WAREHOUSES:
        WAREHOUSES_RELATION[warehouse] = list()
    for warehouse in WAREHOUSES:
        for lot in populated:
            if lot['warehouse'] == warehouse and LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
                WAREHOUSES_RELATION[warehouse].append(lot)
            else:
                WAREHOUSES_RELATION[warehouse].append(lot)
                WAREHOUSES_OUTLOT.append(lot)
    
# def arrangeData():
#     global BUYERS
#     for lot in populated:
#         if lot['buyer_code'] != '': BUYERS.append(lot['buyer_code'])
#     BUYERS = set(BUYERS)
#     for buyer in BUYERS:
#         BUYERS_RELATION[buyer] = list()
#     for buyer in BUYERS:
#         for lot in populated:
#             if lot['buyer_code'] == buyer and LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
#                 BUYERS_RELATION[buyer].append(lot)
#             else:
#                 BUYERS_OUTLOT.append(lot)
# arrangeData()

def formatAddress(address):
    def splitAddress(val):
        data = re.split("[,]+", str(val), flags=re.IGNORECASE)
        if(data[1][0] == ' '):
            data[1] = re.sub(' ', '', data[1], 1)
        return data
    if(address !=  None):
        if(re.search(',', address)):
            return splitAddress(address)
        else:
            counter = 0
            if(re.search('Mombasa', address)):
                address = address.replace('Mombasa', ',Mombasa')
            elif(re.search('Nairobi', address)):
                address = address.replace('Nairobi', ',Nairobi')
            elif(re.search('Kericho', address)):
                address = address.replace('Kericho', ',Kericho')
            else:
                address += ',Nairobi'
            return splitAddress(address)
    else:
        return ['', '']
    
def populate_number(val, level):
    data_length = len(val)
    functions = [
        'SUM', 'PRODUCT', 'DIFFERENCE', 'AVG',
    ]
    hasfn = False
    for fn in functions:
        if(re.search(fn, val)):
            hasfn = True
    counter = 0
    brack = False
    for v in val:
        if(v == '('):
            brack = True
        if v in DATA_LETTERS:
            if(not hasfn):
                val = val.replace(v, v+str(level))
            else:
                if(brack == True):
                    val = val.replace(v, v+str(level))
        if(counter == data_length-1):
            return val
        counter += 1

NUMBER_FORMAT_CELLS = list()
def PopulateRow(sheet, level, row_data, catalogue_data):
    global NUMBER_FORMAT_CELLS
    NUMBER_FORMAT_CELLS = list()
    for data in DATA_SALE:
        # mark = row_data['mark']
        # warehouse = DatabaseQueryProducerCompany(mark)
        # mark = None
        if(data[0] != '_'):
            # if(data == 'warehouse'):
            #     sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = warehouse
            if(data == 'warehouse'):
                # print('row data')
                # print(row_data)
                # print(row_data['invoice_number_buyer'])
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = GetInvoiceWarehouse(catalogue_data, row_data['invoice_number_buyer'])
            elif(data == 'packages' or data == 'net'):
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = int(row_data[data])
            else:
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = row_data[data]
            Format.GeneralCenter(sheet[str(str(DATA_SALE_RELATION[data])+str(level))])
        else:
            sheet[populate_number(DATA_SALE_RELATION[data], level)] = populate_number(DATA_SALE[data], level)
            sheet[populate_number(DATA_SALE_RELATION[data], level)].number_format = '$#,##0.00'
            Format.GeneralCenter(sheet[populate_number(DATA_SALE_RELATION[data], level)])
            cell = populate_number(DATA_SALE_RELATION[data], level)
            NUMBER_FORMAT_CELLS.append(cell)
    return NUMBER_FORMAT_CELLS

def GetInvoiceWarehouse(catalogue_data, invoice_number):
    lot_warehouse = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(data['invoice_number'] == invoice_number):
                    lot_warehouse = data['warehouse']
    return lot_warehouse

def GenerateAccountSale(data, custom_values, counter, type="default"):
    folder='media/resources/'
    fs = FileSystemStorage(location=folder)
    template_default = fs.open('ACCOUNT SALE TEMPLATE.xlsx', 'rb+')
    file_data = template_default
    
    ACsale_template = load_workbook(filename = file_data)
    account_sale_values = GenerateAccountSaleNumber(custom_values['auction_number_alt'], custom_values['auction_number_0'])
    account_sale_number = account_sale_values['number']
    account_sale_number_alt = account_sale_values['number_alt']
    sale_date = custom_values['sale_date']
    prompt_date = custom_values['prompt_date']
    
    catalogue_data = custom_values['catalogue_data']
    invoice_data = custom_values['account_sale_data']
    folder='media/documents/catalogue_data'
    fsc = FileSystemStorage(location=folder)
    
    with fsc.open(catalogue_data, 'rb+') as fcc_file:
        file_datac = json.load(fcc_file)

    # print(file_datac)

    invoice_file = GenerateAccountSaleNumber(custom_values['auction_number'], custom_values['auction_number_0'])['file']

    fs_save_folder = 'media/documents/account_sales/'
    _filename = 'AccountSale_' + invoice_file + '.xlsx'
    dest_save_path = fs_save_folder + _filename

    for warehouse in data:
        warehouse_search = warehouse
        if re.search('CTC', warehouse):
            warehouse_search = 'CTC'
        warehouse_company = DatabaseQueryWarehouseCompany(warehouse)
        warehouse_address = DatabaseQueryWarehouseAddress(warehouse_search)
        code = warehouse
        address_line1 = warehouse_company
        address_line2 = formatAddress(warehouse_address)[0]
        address_line3 = formatAddress(warehouse_address)[1]
        meta_relation = {
            'account_sale_number': account_sale_number_alt,
            'sale_date':  sale_date,
            'sale_date_alt': sale_date,
            'prompt_date': prompt_date,
            'receiver_address_line1': address_line1,
            'receiver_address_line2': address_line2,
            'receiver_address_line3': address_line3,
            'auction_number': custom_values['auction_number'],
            'auction_number_alt': custom_values['auction_number_alt']
        }
        lot_main = list()
        lot_secondary = list()
        lot_start = 15
        lot_secondary_start = 17
        lot_limit_start = lot_start
        totals = 19
        tax_summary = 23
        lot = data[warehouse]
        
        for sale in lot:
            for value in sale:
                if value == 'grade':
                    if TEAGRADES_DATA[sale[value]] == 'primary':
                        lot_main.append(sale)
                    else:
                        lot_secondary.append(sale)
        
        main_length = len(lot_main)
        secondary_length = len(lot_secondary)
        data_length = len(lot)+2
        lot_secondary_ac_start = main_length+lot_secondary_start
        totals += data_length
        tax_summary += data_length
        if(data_length > 1):
            ACsale_template.active.insert_rows(37, data_length-1)
            ACsale_template.active.insert_rows(lot_start, main_length-1)
            ACsale_template.active.insert_rows(lot_secondary_ac_start, secondary_length-1)
        NUMBER_CELLS = list()
        for lot_data in lot_main:
            NUMBER_CELLS = [*NUMBER_CELLS, *PopulateRow(ACsale_template.active, lot_start, lot_data, file_datac)]
            lot_start += 1
        for lot_data in lot_secondary:
            NUMBER_CELLS = [*NUMBER_CELLS, *PopulateRow(ACsale_template.active, lot_secondary_ac_start, lot_data, file_datac)]
            lot_secondary_ac_start += 1
        for cell in NUMBER_CELLS:
            if re.search('J', cell):
                ACsale_template.active[cell].number_format = '0.00_);(0.00)'
            else:
                ACsale_template.active[cell].number_format = '#,##0.00'
        lot_end = lot_start-1
        SUMMARY_RELATION = {}
        for total in DATA_ACCOUNT_SALE_TOTALS:
            ACsale_template.active[DATA_ACCOUNT_SALE_TOTALS[total]+str(totals)] = '=SUM(' + DATA_ACCOUNT_SALE_TOTALS[total] + str(lot_limit_start) + ':' + DATA_ACCOUNT_SALE_TOTALS[total] + str(lot_end) + ')'
            SUMMARY_RELATION[total] = '=' + DATA_ACCOUNT_SALE_TOTALS[total]+str(totals)
        for summary in DATA_ACCOUNT_SALE_TAX_SUMMARY:
            if(summary != 'gross'):
                ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)] = SUMMARY_RELATION[summary]
                ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '$#,##0.00'
            else:
                ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)] = '=SUM(' + DATA_ACCOUNT_SALE_TAX_SUMMARY['amount'] + str(tax_summary) + ':' + DATA_ACCOUNT_SALE_TAX_SUMMARY['brokerage'] + str(tax_summary) + ')'
                ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '$#,##0.00'
        for meta in DATA_ACCOUNT_SALE_META:
            ACsale_template.active[DATA_ACCOUNT_SALE_META[meta]] = meta_relation[meta]

        ACsale_template.active.merge_cells('A' + str(tax_summary) + ':B' + str(tax_summary))
        ACsale_template.active.merge_cells('C' + str(tax_summary) + ':D' + str(tax_summary))
        ACsale_template.active.merge_cells('E' + str(tax_summary) + ':F' + str(tax_summary))
        ACsale_template.active.merge_cells('G' + str(tax_summary) + ':H' + str(tax_summary))
        ACsale_template.active.merge_cells('I' + str(tax_summary) + ':J' + str(tax_summary))
        ACsale_template.active.merge_cells('A' + str(tax_summary-1) + ':B' + str(tax_summary-1))
        ACsale_template.active.merge_cells('C' + str(tax_summary-1) + ':D' + str(tax_summary-1))
        ACsale_template.active.merge_cells('E' + str(tax_summary-1) + ':F' + str(tax_summary-1))
        ACsale_template.active.merge_cells('G' + str(tax_summary-1) + ':H' + str(tax_summary-1))
        ACsale_template.active.merge_cells('I' + str(tax_summary-1) + ':J' + str(tax_summary-1))
        
        ACsale_template.active.title = invoice_file
        
        ACsale_template.save(filename=dest_save_path)
        
        return _filename

class PopulateInvoice():
    def fill_lots(custom_values):
        counter = int(custom_values['account_sale_number'])
        dirs = list()
        for warehouse in WAREHOUSES_RELATION:
            dirs.append(
                GenerateAccountSale({
                    warehouse: WAREHOUSES_RELATION[warehouse]
                }, custom_values, counter, 'default')
            )
            counter += 1
        CloseAccountSaleNumber(counter, custom_values['auction_Pid'])
        return {
            'dirs': dirs
        }

def ACCOUNTSALEGENERATOR(input_data, custom_data):
    StackGenerator(input_data, custom_data['catalogue_data'])
    return PopulateInvoice.fill_lots(custom_data)
